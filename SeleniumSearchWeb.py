import time
from selenium import webdriver
import openpyxl
import datetime


try:

    workbook = openpyxl.load_workbook(
        'C:/Users/ssss/New folder/Automated-Keyword-Search-and-Collect-Google-Suggetions/Excel.xlsx')

except Exception as e:
    print(f"Error loading workbook: {e}")
    exit()


today = datetime.datetime.now()
day_name = today.strftime("%A")

all_sheets = workbook.sheetnames
current_sheet = None
for name in all_sheets:
    if name == day_name:
        current_sheet = workbook[name]


terms_to_search = []
row_numbers = []
for row in range(2, 13):
    cell_value = current_sheet.cell(row=row, column=3).value
    if cell_value:
        terms_to_search.append(cell_value)
        row_numbers.append(row)


browser = webdriver.Chrome()
browser.get("https://www.google.com")

counter = 0
flag = 0

while counter < len(terms_to_search):
    current_term = terms_to_search[counter]
    current_row = row_numbers[counter]

    try:

        time.sleep(2)
        search_input = browser.find_element("name", "q")

        search_input.clear()
        time.sleep(0.5)
        search_input.send_keys(current_term)
        time.sleep(2)

        all_suggestions = []
        suggestion_elements = browser.find_elements(
            "xpath", "//ul[@role='listbox']//li//span")

        for element in suggestion_elements:
            text = element.text
            if text.strip() != "":
                all_suggestions.append(text)

        long = " "
        short = " "

        if len(all_suggestions) > 0:
            long = all_suggestions[0]
            short = all_suggestions[0]

            for suggestion in all_suggestions:
                if len(suggestion) > len(long):
                    long = suggestion
                if len(suggestion) < len(short):
                    short = suggestion

        current_sheet.cell(row=current_row, column=4).value = long
        current_sheet.cell(row=current_row, column=5).value = short

        workbook.save(
            'C:/Users/ssss/New folder/Automated-Keyword-Search-and-Collect-Google-Suggetions/Excel.xlsx')

    except Exception as e:
        print("Something went wrong:", str(e))
        flag = 10

    counter = counter + 1

if flag == 0:
    print("Suggestions stored succesfully in the Excel file")

time.sleep(2)
browser.quit()
